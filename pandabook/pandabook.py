
import pandas as pd
import copy

from typing import Optional, Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pandabook.styles.defaults import DEFAULT_TABLE_STYLE, DEFAULT_HEADING_STYLE, DEFAULT_HEADING_PK_STYLE, DEFAULT_HEADING_HEIGHT, DEFAULT_COLUMN_PK_STYLE, SHRINK_TO_FIT

from functools import lru_cache
from pandabook.utils.sanitise import sanitise_value
from pandabook.utils.format_by_value import format_by_value



# list
# dict
# None / na
# uuid
# hyperlink


# workbook
#     title
#     description
#     author
#     template_info
#     url
#     add_coverpage
#     add_index
#     default_table_style

#     tables
#         df
#         name
#         description
#         autofit_columns
#         start_row
#         start_col
#         style

#         columns
#             description
#             is_pk
#             width
#             style
#             number_format
#             conditional_styles





def get_excel_table(filename, table_name, convert_nan_to_none=True, wb=None):

    if not wb:
        wb = load_workbook(filename, data_only=True)

    for ws in wb.worksheets:
        for entry, data_boundary in ws.tables.items():
            
            if entry != table_name:
                continue

            # Parse the data within the ref boundary
            data = ws[data_boundary]
            # The inner list comprehension gets the values for each cell in the table
            content = [[_strip_text(cell.value) for cell in ent] for ent in data]
            header = content[0]
            rest = content[1:]
            df = pd.DataFrame(rest, columns=header)

            if convert_nan_to_none:
                # pandas NaN dataype can be annoying and cause type issues
                df = df.astype(object).where(df.notnull(), None)

            return df
        

def get_excel_tables(filename, convert_nan_to_none=True):

    wb = load_workbook(filename, data_only=True)
    
    mapping = {}

    for ws in wb.worksheets:
        for table_name in ws.tables.keys():
            df = get_excel_table(filename, table_name, convert_nan_to_none=convert_nan_to_none, wb=wb)
            mapping[table_name] = df

    return mapping

        

def _strip_text(value):
    # Remove leading and trailing whitespace which is a common problem
    try:
        return value.strip()
    except:
        return value






class PandaBook():

    def __init__(
            self,
            title: str = None,
            description: str = None,
            author: str = None,
            url: str = None,
            add_coverpage: bool = True,
            add_index: bool = True,
            use_numbered_sheets: bool = True,
            default_table_style : Optional[TableStyleInfo] = DEFAULT_TABLE_STYLE,
            tables=None
    ):
        
        self.title = title
        self.description = description
        self.author = author
        self.url = url
        self.add_coverpage = add_coverpage
        self.add_index = add_index
        self.use_numbered_sheets = use_numbered_sheets

        self.default_table_style = default_table_style

        self.wb = Workbook(write_only=False)
        first_worksheet = self.wb.worksheets[0]
        self.wb.remove(first_worksheet)

        self.sheets = []
        self.index_page = None

        if tables:
            for name, data in tables.items():
                df = data.get('df')
                name = data.get('name')
                description = data.get('description')
                start_row = data.get('start_row')
                start_column = data.get('start_row')
                index = data.get('index')
                columns = data.get('columns')

                self.add_table(
                    df=df,
                    name=name,
                    description=description,
                    start_row=start_row,
                    start_column=start_column,
                    index=index,
                    columns=columns
                )

    def add_table(
        self,
        df: pd.DataFrame,
        name: str,
        description: Optional[str] = None,
        start_row: Optional[int] = 1,
        start_column: Optional[int] = 1,
        index: bool = False,
        sheet_position: int = None,
        exclude_from_index: bool = False,
        columns: Dict[str, Dict[str , object]] = None,
    ):

        panda_sheet = PandaSheet(
            panda_book=self,
            df=df,
            name=name,
            description=description,
            start_row=start_row,
            start_column=start_column,
            index=index,
            sheet_position=sheet_position,
            columns=columns,
        )
        if not exclude_from_index:
            self.sheets.append(panda_sheet)

        return panda_sheet
    
    def save(self, *args, **kwargs):

        if self.add_index:
            self._add_table_summary()
        if self.add_coverpage:
            pass
            # TODO self._add_coverpage()

        self.wb.save(*args, **kwargs)


    def _add_table_summary(self):
        
        summary_data = [panda_sheet.get_summary() for panda_sheet in self.sheets]
        table_summary_df = pd.DataFrame(summary_data)
        table_summary_df['link'] = None

        # table_summary_df['sheet_name'] = table_summary_df['sheet_name'].apply(lambda x : f'=HYPERLINK("#\'{x}\'!A1", "{x}")')
        columns = {
            'sheet_name': {'width' : 15},
            'table_name': {'width' : 20},
            'description': {'width' : 60},
            'num_records' : {'width' : 15},
            'link': {'width' : 10},
        }
        self.index_page = self.add_table(
            df=table_summary_df,
            name='index',
            description='Table of contents',
            index=False,
            columns=columns,
            exclude_from_index=True,
            sheet_position=0
        )
        
        ws = self.wb['index']
        
        for row_index, sheet_name in enumerate(table_summary_df['sheet_name']):
            sheet_row_no = row_index + 2
            cell = ws.cell(column=5, row=sheet_row_no)
            cell.value = f"""=HYPERLINK("#\'{sheet_name}\'!A1", "{sheet_name}")"""
            cell.style = 'Hyperlink'





class PandaSheet():

    def __init__(
        self,
        panda_book: PandaBook,
        df: pd.DataFrame,
        name: str,
        description: Optional[str] = None,
        start_row: Optional[int] = 1,
        start_column: Optional[int] = 1,
        index: bool = False,
        autofit_columns: bool = True,
        sheet_position: Optional[int] = None,
        columns: Dict[str, Dict[str, object]] = None
    ):
        
        self.panda_book = panda_book
        self.df = df
        self.name = name
        self.description = description
        self.start_row = start_row
        self.start_column = start_column
        self.index = index
        self.autofit_columns = autofit_columns
        self.sheet_position = sheet_position

        if self.panda_book.use_numbered_sheets and name not in ['index', 'cover']:
            num_sheets = len(self.panda_book.sheets) + 1
            self.sheet_name = f'{num_sheets:02d}'
        else:
            if len(name) > 31:
                raise ValueError('The table name {name} exceeds 31 characters. Shorten the name, or use numbered sheets.')
            self.sheet_name = name




        if not columns:
            self.columns = {column_name : {} for column_name in self.df.columns}
        else:
            self.columns = columns

        self._add_worksheet()
        self._write_column_header()
        self._write_data()
        self._appply_formatting()
        self._add_named_table()

        if self.autofit_columns:
            self._autofit_columns()



    def get_summary(self):
        data = {
            'sheet_name' : self.sheet_name,
            'table_name' : self.name,
            'description' : self.description,
            'num_records' : len(self.df)
        }
        return data



    def _add_worksheet(self):
        try:
            # Delete the sheet if it already exists
            del self.panda_book.wb[self.sheet_name]
        except:
            pass
        ws = self.panda_book.wb.create_sheet(title=self.sheet_name, index=self.sheet_position)
        self.ws = ws

    @lru_cache
    def _get_header_values(self):
        ws = self.ws
        df = self.df

        headers = list(df.columns)

        if self.index:
            headers = [
                name if name is not None else f'level_{idx}' if df.index.nlevels > 1 else 'index'
                for idx, name in enumerate(df.index.names)
            ] + headers
        
        headers = [str(header) for header in headers]

        df.reset_index
        return headers

    @lru_cache
    def _is_pk(self, column_name):
        is_pk = False
        if self.index:
            headers = self._get_header_values()
            header_index = headers.index(column_name)
            is_pk = header_index  < self.df.index.nlevels

        is_pk = self.columns.get(column_name, {}).get('is_pk', is_pk)
        
        return is_pk

        
    def _write_column_header(self):

        ws = self.ws
        df = self.df

        headers = self._get_header_values()

        for sheet_column_no, column_name in enumerate(headers, start=self.start_column):

            cell = ws.cell(column=(sheet_column_no), row=self.start_row)
            cell.value = column_name

            description = self.columns.get(column_name, {}).get('description')
            
            if description:
                cell.comment = Comment(description)

            if self._is_pk(column_name):
                cell.style = DEFAULT_HEADING_PK_STYLE
            else:
                cell.style = DEFAULT_HEADING_STYLE

            column_width = self.columns.get(column_name, {}).get('width')

            if column_width:
                ws.column_dimensions[
                    get_column_letter(sheet_column_no)
                ].width = column_width

        ws.row_dimensions[self.start_row].height = DEFAULT_HEADING_HEIGHT


    def _write_data(self):
        df = self.df
        ws = self.ws

        for row in df.itertuples(index=self.index):

            if isinstance(row[0], tuple):
                # Unpack multi-index if applicable
                row = row[0] + row[1:]
            
            row = tuple(sanitise_value(x) for x in row)

            left_padding = self.start_column - 1
            row_to_write = (None,) * left_padding + row
            ws.append(row_to_write)

    def _appply_formatting(self):

        headers = self._get_header_values()
        drop = not self.index
        df = self.df.reset_index(drop=drop)
        df.columns = headers


        for header_index, column_name in enumerate(headers):
            
            # # Using index rather than column names to account for multi-indexed columns
            # column_series = df.iloc[:,header_index]
            series = df[column_name]
            
            sheet_column_no = self.start_column + header_index
            column_settings = self.columns.get(column_name, {})
            
            is_pk = self._is_pk(column_name)

            column_style = column_settings.get('style')
            column_number_format = column_settings.get('number_format')
            
            conditional_style = column_settings.get('conditional_style', {})
            
            for sheet_row_no, value in enumerate(series, self.start_row + 1):
                cell_style = format_by_value(value=value)
                if is_pk:
                    cell_style = DEFAULT_COLUMN_PK_STYLE # Take PK style over others
                cell_number_format = None
                try:
                    cell_conditional_style = conditional_style.get(value)
                except:
                    # wont work on unhashable values
                    pass

                cell_style = cell_conditional_style or cell_style or column_style
                cell_number_format = cell_number_format or column_number_format
                

                

                if cell_style or cell_number_format:
                    cell = self.ws.cell(column=sheet_column_no, row=sheet_row_no)
                    if cell_style:
                        cell.style = cell_style

                    if cell_number_format:
                        cell.number_format = cell_number_format



    def _add_named_table(self):
        
        ws = self.ws
        df = self.df
        number_of_columns = len(df.columns)

        if self.index:
            number_of_columns += df.index.nlevels

        table_range = CellRange(
            min_col=self.start_column,
            min_row=self.start_row,
            max_col=self.start_column + number_of_columns - 1,
            max_row=self.start_row + max(len(df), 1), # Need at least one row for an Excel table. A blank dataframe will have one blank row.
        )

        table = Table(displayName=self.name, ref=table_range.coord)

        table_style = self.panda_book.default_table_style
        table.tableStyleInfo = table_style

        ws.add_table(table)
        ws.freeze_panes = f"A{self.start_row + 1}"

    def _get_conditional_style(self, value, conditional_styles):

        return conditional_styles.get(value)

    
    def _autofit_columns(self, max_allowable=80):
        for col in self.ws.columns:

            column_name = col[self.start_row - 1].value
            column = get_column_letter(col[0].column)  # Get the column letter

            column_width = self.columns.get(column_name, {}).get('width')
            if column_width:
                self.ws.column_dimensions[column].width = column_width
                continue

            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column letter

            for cell in col:
                value = cell.value
                if value:
                    cell_length = len(str(value))
                    if cell_length > max_length:
                        max_length = cell_length
                    if cell_length > max_allowable:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')


            self.ws.column_dimensions[column].width = (
                min(max_length, max_allowable) + 2 * 1.1
            )


 ####################### TODO


    @staticmethod
    def get_configuration_template(data):
        
        if type(data) is pd.DataFrame:
            data = [data]

        configuration = {
            'title' : None,
            'description' : None,
            'author' : None,
            'url' : None,
            'add_coverpage' : True,
            'add_index' : True,
            'default_table_style' : str(DEFAULT_TABLE_STYLE.name),
            'tables' : {}
        }

        for index, table in enumerate(data):
            table_name = f'table{index:0d}'
            table_data = {
                'df' : f'df{index:0d}',
                'description' : 'Placeholder',
                'start_row' : 1,
                'start_column' : 1,
                'columns' : {}
            }
            for col_name in table.columns:
                column_data = {
                    'description' : 'Placeholder',
                    'width' : None,
                    'auto_width' : True,

                }
                table_data['columns']['col_name'] = column_data

            configuration['tables'][table_name] = table_data
            return configuration