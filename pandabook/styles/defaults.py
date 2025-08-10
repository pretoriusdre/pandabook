# from openpyxl import load_workbook, Workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.worksheet.cell_range import CellRange

# from openpyxl.comments import Comment
# from openpyxl.utils import get_column_letter
# from openpyxl.styles.borders import Border, Side
# from openpyxl.styles.fills import PatternFill 


from openpyxl.styles import NamedStyle, Alignment, PatternFill
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.styles import Font, Color
from openpyxl.styles import colors


DEFAULT_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

white = Font(color=colors.WHITE)

blue = '4F81BD'
blue_darker = '3E6594'

DEFAULT_HEADING_STYLE = NamedStyle(name="header")
DEFAULT_HEADING_STYLE.alignment = Alignment(vertical="top")
DEFAULT_HEADING_STYLE.fill = PatternFill(fill_type="solid", fgColor=blue)
DEFAULT_HEADING_STYLE.font = white

DEFAULT_HEADING_PK_STYLE = NamedStyle(name="header_pk")
DEFAULT_HEADING_PK_STYLE.alignment = Alignment(vertical="top")
DEFAULT_HEADING_PK_STYLE.fill = PatternFill(fill_type="solid", fgColor=blue_darker)
DEFAULT_HEADING_PK_STYLE.font = white


DEFAULT_HEADING_HEIGHT = 32


DEFAULT_COLUMN_PK_STYLE = NamedStyle(name="pk")
DEFAULT_COLUMN_PK_STYLE.fill = PatternFill(fill_type="solid", fgColor=blue_darker)
DEFAULT_COLUMN_PK_STYLE.alignment = Alignment(shrink_to_fit=True) # For uuid type pk this is useful
DEFAULT_COLUMN_PK_STYLE.font = white



SHRINK_TO_FIT = NamedStyle(name="shrink")
SHRINK_TO_FIT.alignment = Alignment(shrink_to_fit=True)


DATE_ISO_STYLE = NamedStyle(name="date")
DATE_ISO_STYLE.number_format = "yyyy-mm-dd"
DATE_ISO_STYLE.alignment = Alignment(shrink_to_fit=True)

DATETIME_ISO_STYLE = NamedStyle(name="datetime_iso")
DATETIME_ISO_STYLE.number_format = 'yyyy-mm-dd"T"hh:mm:ss'
DATETIME_ISO_STYLE.alignment = Alignment(shrink_to_fit=True)